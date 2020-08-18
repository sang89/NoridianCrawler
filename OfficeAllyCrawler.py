import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent

from openpyxl import load_workbook

from utils import *
from PatientInfo import *
from ExcelUtils import *
from constants import *
from NoridianCrawler import WebTable

options = Options()
ua = UserAgent()
userAgent = ua.random
print(userAgent)
options.add_argument('user-agent={userAgent}')
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches",["enable-automation"])

driver = webdriver.Chrome(options=options, executable_path='D:\ChromDriver\chromedriver.exe')
url = driver.command_executor._url
print(url)
session_id = driver.session_id
print(session_id)

class OfficeAllyTable():
    def __init__(self, webtable_path):
        self.path = webtable_path
        self.row_offset = 1
        self.column_offset = 5
        self.detail_column = 15

    def get_row_count(self):
        return len(driver.find_elements_by_xpath(self.path + "/tbody/tr")) - 1

    def get_column_count(self):
       #return len(driver.find_elements_by_xpath(self.path + '/tbody/tr[1]/td'))
        return 14

    def get_table_size(self):
        return {"rows": self.get_row_count(),
                "columns": self.get_column_count()}

    def get_all_data(self):
        # get number of rows
        noOfRows = self.get_row_count()
        # get number of columns
        noOfColumns = self.get_column_count()
        allData = []
        # iterate over the rows, to ignore the headers we have started the i with '1'
        for i in range(1, noOfRows + 1):
            # reset the row data every time
            ro = []
            # iterate over columns
            for j in range(1, noOfColumns + 1):
                # get text from the i th row and j th column
                # sometime the row is empty, we just append an empty string
                try:
                    ro.append(
                        driver.find_element_by_xpath(
                            self.path +
                            "/tbody/tr[" + str(self.row_offset+i) + "]/td[" + str(self.column_offset+j) + "]").text)
                except:
                    ro.append('')
            # add the row data to allData of the self.table
            allData.append(ro)

        return allData

    def row_data(self, row_number):
        row = driver.find_elements_by_xpath(self.path + "/tbody/tr["+str(self.row_offset + row_number)+"]/td")
        rData = []
        for webElement in row :
            rData.append(webElement.text)
        return rData

    def column_data(self, column_number):
        col = driver.find_elements_by_xpath(self.path + "/tbody/tr/td["+str(self.column_offset + column_number)+"]")
        rData = []
        for webElement in col :
            rData.append(webElement.text)
        return rData

    def get_patient_info(self, row_number):
        #[lastname, first name, dob]
        row_data = self.row_data(row_number)
        if len(row_data) > 0:
            return [row_data[5], row_data[6], row_data[8]]
        return []

    def go_to_detail(self, row_number):
        current_cell = driver.find_element_by_xpath(self.path + "/tbody/tr[" + str(self.row_offset + row_number) +
                                                    "]/td[" + str(self.detail_column) + "]")
        link_to_click = current_cell.find_element_by_tag_name("a")
        link_to_click.click()
        time.sleep(2)
        result = self.get_detail_info()
        return result

    def get_detail_info(self):
        last_name = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucPatient_LastName"]').get_attribute('value')
        first_name = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucPatient_FirstName"]').get_attribute('value')
        month = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucPatient_DOB_Month"]').get_attribute('value')
        day = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucPatient_DOB_Day"]').get_attribute('value')
        year = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucPatient_DOB_Year"]').get_attribute('value')
        dob = convert_to_datetime_obj(month+'/'+day+'/'+year).strftime('%m/%d/%Y')
        insurance_tab = driver.find_element_by_xpath('//*[@id="PatientTabs"]/ul/li[2]/a')
        insurance_tab.click()
        time.sleep(1)
        medicare_number = driver.find_element_by_xpath(
            '//*[@id="ctl00_phFolderContent_ucPatient_InsuranceSubscriberID"]').get_attribute('value')
        if not medicare_number:
            medicare_number = ''
        patient_chart_tab = driver.find_element_by_xpath('//*[@id="patient-charts_tab"]')
        patient_chart_tab.click()
        time.sleep(2)
        return [last_name, first_name, medicare_number, dob]

    def look_up_patient_info(self, first_name, last_name, dob, medicare_number):
        # first look for the applicable row
        num_of_row = self.get_row_count()
        result_row_num = -1
        for i in range(1, num_of_row + 1):
            current_row_data = self.get_patient_info(i)
            print('current_row_data', current_row_data)
            if (last_name == current_row_data[0] and first_name == current_row_data[1]) or \
                    convert_to_datetime_obj(current_row_data[2]) == dob:
                result_row_num = i
                break

        if result_row_num > 0:
            print('result_row_num', result_row_num)
            result = self.go_to_detail(result_row_num)
            print('go_to_detail result', result)
            return result
        else:
            print('No applicable row!')
            return []

def main():
    list_of_queries = get_list_of_queries()

    book = load_workbook(SOURCE_CLEAN_FILE_NAME)
    with pd.ExcelWriter(SOURCE_CLEAN_FILE_NAME, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for patient_cntr in range(len(list_of_queries)):
        #for patient_cntr in range(1):
            list_of_columns = ['LAST', 'FIRST', 'MEDICARE #', 'DOB', 'From DOS', 'To Dos', 'Status']
            df = pd.DataFrame(columns=list_of_columns)

            # enter the info
            cur_info = list_of_queries[patient_cntr]
            first_name = cur_info.first_name.upper()
            last_name = cur_info.last_name
            medicare_number = cur_info.medicare_number
            dob = convert_to_datetime_obj(cur_info.dob)
            from_dos = cur_info.from_dos
            to_dos = cur_info.to_dos

            patientLastNameOption = driver.find_element_by_xpath(
                '//*[@id="ctl00_phFolderContent_ucSearch_lstSearchBy"]/option[16]').click()
            lastNameOptionInput = driver.find_element_by_xpath(
                '//*[@id="ctl00_phFolderContent_ucSearch_txtSearch"]'
            )
            lastNameOptionInput.click()
            lastNameOptionInput.clear()
            lastNameOptionInput.send_keys(last_name)
            time.sleep(1)
            submitBtn = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucSearch_btnSearch"]')
            submitBtn.click()
            time.sleep(3)
            table = OfficeAllyTable('//*[@id="ctl00_phFolderContent_myCustomGrid_myGrid"]')
            print('Investigating patient:')
            print('first name', first_name)
            print('last name', last_name)
            print('dob', dob)
            print('medicare', medicare_number)
            result = table.look_up_patient_info(first_name, last_name, dob, medicare_number)

            # if cannot find anything, try reversing first, last name
            if len(result) == 0:
                driver.find_element_by_xpath(
                    '//*[@id="ctl00_phFolderContent_ucSearch_lstSearchBy"]/option[16]').click()
                lastNameOptionInput = driver.find_element_by_xpath(
                    '//*[@id="ctl00_phFolderContent_ucSearch_txtSearch"]'
                )
                first_name, last_name = last_name, first_name
                lastNameOptionInput.click()
                lastNameOptionInput.clear()
                lastNameOptionInput.send_keys(last_name)
                time.sleep(1)
                submitBtn = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucSearch_btnSearch"]')
                submitBtn.click()
                time.sleep(3)
                table = OfficeAllyTable('//*[@id="ctl00_phFolderContent_myCustomGrid_myGrid"]')
                print('Investigating patient with first and last names reversed:')
                print('first name', first_name)
                print('last name', last_name)
                print('dob', dob)
                print('medicare', medicare_number)
                result = table.look_up_patient_info(first_name, last_name, dob, medicare_number)

            # if still cannot find anything, try first name
            if len(result) == 0:
                driver.find_element_by_xpath(
                    '//*[@id="ctl00_phFolderContent_ucSearch_lstSearchBy"]/option[14]').click()
                first_name_input = driver.find_element_by_xpath(
                    '//*[@id="ctl00_phFolderContent_ucSearch_txtSearch"]'
                )
                first_name_input.click()
                first_name_input.clear()
                first_name_input.send_keys(first_name)
                time.sleep(1)
                submitBtn = driver.find_element_by_xpath('//*[@id="ctl00_phFolderContent_ucSearch_btnSearch"]')
                submitBtn.click()
                time.sleep(3)
                table = OfficeAllyTable('//*[@id="ctl00_phFolderContent_myCustomGrid_myGrid"]')
                print('Investigating patient with first name:')
                print('first name', first_name)
                print('last name', last_name)
                print('dob', dob)
                print('medicare', medicare_number)
                result = table.look_up_patient_info(first_name, last_name, dob, medicare_number)

            if len(result) > 0:
                result = result + [from_dos, to_dos, 1]
            else:
                # just append the existing data
                result = [last_name, first_name, medicare_number, dob.strftime('%m/%d/%Y'), from_dos, to_dos, 0]

            row_to_add = pd.Series(result, index=df.columns)
            df = df.append(row_to_add, ignore_index=True)

            df.to_excel(writer, sheet_name=SOURCE_CLEAN_SHEET_NAME, header=None, startrow=book[SOURCE_CLEAN_SHEET_NAME].max_row)
            writer.save()

if __name__ == '__main__':
    #driver.get('https://pm.officeally.com')
    driver = webdriver.Remote(command_executor='http://127.0.0.1:56287',desired_capabilities={})
    driver.close()   # this prevents the dummy browser
    driver.session_id = '9504382faf47224de4c9f737ed981734'
    main()