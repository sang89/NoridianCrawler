import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook

from utils import *
from PatientInfo import *
from ExcelUtils import *
from constants import *

driver = webdriver.Chrome('D:\ChromDriver\chromedriver.exe')
url = driver.command_executor._url
print(url)
session_id = driver.session_id
print(session_id)

class WebTable:
    def __init__(self, webtable_path):
       self.path = webtable_path

    def get_row_count(self):
        #return len(self.table.find_elements_by_tag_name("tr")) - 1
        return len(driver.find_elements_by_xpath(self.path + "/tbody/tr"))

    def get_column_count(self):
        #return len(self.table.find_elements_by_xpath("//tbody/tr[1]/td"))
        return len(driver.find_elements_by_xpath(self.path + '/tbody/tr[1]/td'))

    def get_table_size(self):
        return {"rows": self.get_row_count(),
                "columns": self.get_column_count()}

    def row_data(self, row_number):
        row = driver.find_elements_by_xpath(self.path + "/tbody/tr["+str(row_number)+"]/td")
        rData = []
        for webElement in row :
            rData.append(webElement.text)
        return rData

    def column_data(self, column_number):
        col = driver.find_elements_by_xpath(self.path + "/tbody/tr/td["+str(column_number)+"]")
        rData = []
        for webElement in col :
            rData.append(webElement.text)
        return rData

    def get_all_data(self):
        # get number of rows
        noOfRows = self.get_row_count()
        # get number of columns
        noOfColumns = self.get_column_count()
        allData = []
        # iterate over the rows, to ignore the headers we have started the i with '1'
        for i in range(1, noOfRows+1):
            # reset the row data every time
            ro = []
            # iterate over columns
            for j in range(1, noOfColumns+1) :
                # get text from the i th row and j th column
                # sometime the row is empty, we just append an empty string
                try:
                    ro.append(driver.find_element_by_xpath(self.path + "/tbody/tr["+str(i)+"]/td["+str(j)+"]").text)
                except:
                    ro.append('')
            # add the row data to allData of the self.table
            allData.append(ro)

        return allData

    def click_on_view_detail(self, row_number):
        view_detail_column_number = 8
        current_cell = driver.find_element_by_xpath(self.path + "/tbody/tr["+str(row_number)+"]/td["+str(view_detail_column_number)+"]")
        link_to_click = current_cell.find_element_by_tag_name("a")
        link_to_click.click()
        time.sleep(10)

    def presence_of_data(self, data):
        # verify the data by getting the size of the element matches based on the text/data passed
        dataSize = len(driver.find_elements_by_xpath(self.path + "/tbody/td[normalize-space(text())='"+data+"']"))
        presence = False
        if(dataSize > 0):
            presence = true
        return presence

    def get_cell_data(self, row_number, column_number):
        if (row_number == 0):
            raise Exception("Row number starts from 1")

        #row_number = row_number+1
        cellData = driver.find_element_by_xpath(self.path + "//tbody/tr["+str(row_number)+"]/td["+str(column_number)+"]").text
        return cellData

def signInToNoridian():
    #driver.get('https://www.noridianmedicareportal.com/web/nmp/home')

    driver.maximize_window()
    driver.find_element(By.CSS_SELECTOR, ".ui-state-focus > .ui-button-text").click()
    driver.find_element(By.ID, "username").click()
    driver.find_element(By.ID, "username").send_keys(NORIDIAN_USERNAME)
    driver.find_element(By.ID, "password").click()
    driver.find_element(By.ID, "password").send_keys(NORIDIAN_PASSWORD)
    driver.find_element(By.ID, "btnSubmit").click()
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR, ".m-top-10:nth-child(3) span:nth-child(2)").click()
    driver.find_element(By.ID, "submit-btn").click()
    time.sleep(3)
    driver.find_element(By.ID, "EMailPassword").click()
    time.sleep(20) # the wait here is to enter confirmation code from gmail
    print('We are at email confirm')
    driver.find_element(By.CSS_SELECTOR, "#layout_19").click()
    time.sleep(5)
    print('We are at noridian')

# Open a new window to get to gmail
#driver.execute_script("window.open('https://www.google.com');")
#driver.switch_to.window(driver.window_handles[1])
#driver.find_element(By.LINK_TEXT, "Gmail").click()

def enter_info(patient_info):
    medicare_number = patient_info.medicare_number
    first_name = patient_info.first_name
    last_name = patient_info.last_name
    dob = patient_info.dob
    from_dos = patient_info.from_dos
    to_dos = patient_info.to_dos

    driver.find_element(By.ID, "hicn").click()
    driver.find_element(By.ID, "hicn").clear()
    driver.find_element(By.ID, "hicn").send_keys(medicare_number)
    driver.find_element(By.ID, "firstName").click()
    driver.find_element(By.ID, "firstName").clear()
    driver.find_element(By.ID, "firstName").send_keys(first_name)
    driver.find_element(By.ID, "lastName").click()
    driver.find_element(By.ID, "lastName").clear()
    driver.find_element(By.ID, "lastName").send_keys(last_name)
    driver.find_element(By.ID, "dob").click()
    driver.find_element(By.ID, "dob").clear()
    driver.find_element(By.ID, "dob").send_keys(dob)

    driver.find_element_by_xpath('//*[@id="fromDate"]').click()
    driver.find_element_by_xpath('//*[@id="fromDate"]').clear()
    driver.find_element_by_xpath('//*[@id="fromDate"]').send_keys(from_dos)
    driver.find_element_by_xpath('//*[@id="toDate"]').click()
    driver.find_element_by_xpath('//*[@id="toDate"]').clear()
    driver.find_element_by_xpath('//*[@id="toDate"]').send_keys(to_dos)
    driver.find_element(By.ID, "btnSubmit").click()
    print('Finish submitting info for:')
    print('First name: ', first_name)
    print('Last name: ', last_name)

    #alert = driver.find_element_by_xpath('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[1]/div')
    try:
        alert = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[1]/div')))
        if alert:
            print("Alert:", alert.text)
        else:
            print('Cannot print alert')
    except:
        print('There is no alert')


def get_claim_status_overview(patient_info):
    start = patient_info.from_dos
    end = patient_info.to_dos

    try:
        table = WebDriverWait(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, "//table[@id='_claimstatus_WAR_NMPFrontendportlet__table']")))

        if table:
            w = WebTable("//table[@id='_claimstatus_WAR_NMPFrontendportlet__table']")

            # Find out which row has date of service in the given range
            num_of_rows = w.get_row_count()
            has_applicable_row = num_of_rows > 0
            data_array = []

            if has_applicable_row:
                # go to all row except the last row
                for row_num in range(1, num_of_rows):
                    w.click_on_view_detail(row_num)
                    print('Click on row num: ', row_num)
                    data_array = data_array + process_data()
                    back_btn = driver.find_elements_by_xpath('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[5]/div[2]/form/input[9]')[0]
                    back_btn.submit()

                # go to the last row
                w.click_on_view_detail(num_of_rows)
                print('Click on last row: ', num_of_rows)
                data_array = data_array + process_data()
                return data_array
            else:
                print('There is no applicable row!')
                if table:
                    print('Go back to search!')
                    new_inquiry_btn = driver.find_elements_by_xpath('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[3]/div[2]/a[1]')[0]
                    new_inquiry_btn.click()

                return []
        else:
            return []
    except:
        return []

def process_data():
    diagnosis_pointer_table = WebTable('//*[@id="diagnosiscodeandPositionTable"]/table')
    claim_status_detail_table = WebTable('//*[@id="claim-status-line_details_table"]')
    reason_table = WebTable('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[10]/table')

    # get all the data
    #print('look_up_diagnostic_pointer_table', diagnosis_pointer_table.get_all_data())
    #print('claim_status_detail_table', claim_status_detail_table.get_all_data())
    #print('reason_table', reason_table.get_all_data())

    claim_status_detail_array = claim_status_detail_table.get_all_data()
    num_of_rows = len(claim_status_detail_array)
    for i in range(0, num_of_rows):
        index_of_diagnosis_pointer = 6
        diagnosis_code = claim_status_detail_array[i][index_of_diagnosis_pointer]
        # Now we look up this code at the pointer table
        diagnosis_detail = look_up_diagnostic_pointer_table(diagnosis_code, diagnosis_pointer_table.get_all_data())
        claim_status_detail_array[i][index_of_diagnosis_pointer] = diagnosis_detail

        reasons = claim_status_detail_array[i][-1]
        # Look up this reason in the reason table
        reason_detail = look_up_reason_pointer_table(reasons, reason_table.get_all_data())
        claim_status_detail_array[i][-1] = reason_detail

    print('Claim status detail', claim_status_detail_array)
    return claim_status_detail_array

def get_back_to_new_search():
    back_btn = driver.find_elements_by_xpath('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[5]/div[2]/form/input[9]')[0]
    back_btn.submit()
    new_inquiry_btn = driver.find_elements_by_xpath('//*[@id="p_p_id_claimstatus_WAR_NMPFrontendportlet_"]/div/div/div[3]/div[2]/a[1]')[0]
    new_inquiry_btn.click()

def main():
    # First sign in to noridian
    #signInToNoridian()

    list_of_queries = get_list_of_queries()

    book = load_workbook(MAIN_EXCEL_FILE_NAME)
    with pd.ExcelWriter(MAIN_EXCEL_FILE_NAME, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for patient_cntr in range(len(list_of_queries)):
            list_of_columns = ['FIRST', 'LAST', 'MEDICARE #', 'DOB', 'FIRST IN NOR', 'LAST IN NOR',
                'From DOS', 'To DOS', 'Procedure Code', 'Modifier', 'Units', 'Diagnostic Code', 'Billed Amount',
                'Allowed Amount', 'Provider Paid', 'Reasons']
            df = pd.DataFrame(columns = list_of_columns)

            # enter the info
            cur_info = list_of_queries[patient_cntr]
            first_name = cur_info.first_name
            last_name = cur_info.last_name
            medicare_number = cur_info.medicare_number
            dob = cur_info.dob
            from_dos = cur_info.from_dos
            to_dos = cur_info.to_dos

            done = False
            name_combination = create_name_combinations(first_name, last_name)
            name_cntr = 0
            while not done:
                out_of_combination_option = name_cntr > len(name_combination) - 1
                if (out_of_combination_option):
                    row_array = [first_name, last_name, medicare_number, dob]
                    row_array = row_array + ['', '', from_dos, to_dos]
                    row_array = row_array + ['', '', '', '', '', '', '', '']
                    row_to_add = pd.Series(row_array, index = df.columns)
                    df = df.append(row_to_add, ignore_index=True)
                    break

                cur_name_combination = name_combination[name_cntr]
                new_patient_info = QueryInfo(cur_name_combination[0], cur_name_combination[1], medicare_number, dob, from_dos, to_dos, True)
                enter_info(new_patient_info)
                data_array = get_claim_status_overview(new_patient_info)
                done = len(data_array) > 0
                if not done:
                    name_cntr += 1
                else:
                    for i in range(len(data_array)):
                        row_array = [first_name, last_name, medicare_number, dob]
                        row_array = row_array + [cur_name_combination[0], cur_name_combination[1]]
                        row_array = row_array + data_array[i][1:]
                        row_to_add = pd.Series(row_array, index = df.columns)
                        df = df.append(row_to_add, ignore_index=True)

                    get_back_to_new_search()

            df.to_excel(writer, sheet_name=MAIN_SHEET_NAME, header=None, startrow=book[MAIN_SHEET_NAME].max_row)
            writer.save()


driver.get('https://www.noridianmedicareportal.com/web/nmp/home')
driver = webdriver.Remote(command_executor='http://127.0.0.1:63442',desired_capabilities={})
driver.close()   # this prevents the dummy browser
driver.session_id = 'a10c0f066206b33eede9a3c1b9c1132d'
main()

